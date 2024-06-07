#chargen.py
#classic traveller mass character generator
#version 0.2 - June 7th, 2024
#by Omer Golan-Joel golan2072@gmail.com
#open source under the Creative Commons Zero 1.0 License

import utility
import json
import random
import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import Font

blade_list = ["Blade", "Cutlass", "Sword", "Broadsword", "Polearm", "Cudgel", "Spear", "Dagger"]
gun_list = ["Autopistol", "SMG", "Rifle", "Body Pistol", "Autorifle", "Carbine", "Revolver", "Shotgun", "Laser Carbine", "Laser Rifle"]
vehicle_list = ["Ground Car", "ATV", "AFV", "G-Carrier", "Air/Raft", "Speeder", "Motorboat", "Submersible", "Biplane", "Jump Jet"]

def stat_roll_parser(stat_string):
    stat_name = str(stat_string[:3])
    threshold = stat_string[3]
    stat_list = ["STR", "DEX", "END", "INT", "EDU", "SOC"]
    return [int(stat_list.index(stat_name)), int(threshold)] #returns UPP stat index and threshold


def inventory_check (item, skills):
    if item == "Gun":
        for skill in skills:
            if skill in gun_list:
                item = skill
            else:
                item = random.choice(gun_list)
        else:
            pass
    elif item == "Blade":
        for skill in skills:
            if skill in blade_list:
                item = skill
            else:
                item = random.choice(blade_list)
    elif item == "Vehicle":
        for skill in skills:
            if skill in vehicle_list:
                item = skill
            else:
                item = random.choice(vehicle_list)
    else:
        pass
    return item

def survival(char_data, career, upp):
        survival_roll = utility.dice(2, 6)
        modifier = stat_roll_parser(char_data[career]['survival+2'])
        if int(upp[[modifier][0][0]]) >= int([modifier][0][1]):
            survival_roll += 2
        if survival_roll >= char_data[career]["survival"]:
            survival = True
        else:
            survival = False
        return survival

def commission (char_data, career, upp):
        if char_data[career]['commission+1'] == "":
            return False
        else:
            commission_roll = utility.dice(2, 6)
            modifier = stat_roll_parser(char_data[career]["commission+1"])
            if int(upp[[modifier][0][0]]) >= int([modifier][0][1]):
                commission_roll +=1
            if commission_roll >= int(char_data[career]["commission"]):
                return True
            else:
                return False
            
def promotion (char_data, career, upp, rank):
        if char_data[career]['promotion+1'] == "":
            return False
        else:
            promotion_roll = utility.dice(2, 6)
            modifier = stat_roll_parser(char_data[career]["promotion+1"])
            if int(upp[[modifier][0][0]]) >= int([modifier][0][1]):
                promotion_roll += 1
            if promotion_roll >= int(char_data[career]["promotion"]):
                rank += 1
                return rank, True
            else:
                return rank, False
            

class Character:
    def __init__(self, library_file, death, career_name):
        library = open(library_file) #library_file being a .json file
        self.char_data = json.load(library)
        self.career_list = []
        for career in self.char_data:
            self.career_list.append(career)
        if utility.dice(1,6) <= 3:
            self.name = utility.random_line("Data/malenames.txt") + " " + utility.random_line("Data/surnames.txt")
            self.sex = "Male"
        else:
            self.name = utility.random_line("Data/femalenames.txt") + " " + utility.random_line("Data/surnames.txt")
            self.sex = "Female"
        self.upp = [utility.dice(2, 6), utility.dice(2, 6), utility.dice(2, 6), utility.dice(2, 6), utility.dice(2, 6), utility.dice(2, 6)]
        self.career = career_name
        self.enlistment ()
        self.rank = 0
        self.skills = {}
        self.skill_string = ""
        self.upp_string = ""
        self.inventory = {}
        self.inventory_string = ""
        self.char_string = ""
        self.species = "Human"
        self.status = "Alive"
        self.rank_text = ""
        self.inventory = {}
        i = 0
        self.terms = 0
        self.cash = 0
        self.age = 18
        if "0" in self.char_data[self.career]["autoskills"]:
            self.add_skill(self.char_data[self.career]["autoskills"]["0"])
        while True:
            self.terms += 1
            self.age += 4
            survival_margin = survival(self.char_data, self.career, self.upp)
            if survival_margin == True:
                self.status = "alive"
            elif survival_margin == False and death == True:
                self.status = "DECEASED"
                break
            self.add_skill(self.skill_roll())
            if self.terms == 1:
                self.add_skill(self.skill_roll())
            else:
                pass
            if self.rank == 0:
                commission_check = commission(self.char_data, self.career, self.upp)
                if commission_check == True:
                    self.rank = 1
                    self.add_skill(self.skill_roll())
                else:
                    self.rank = 0
            elif self.rank in range (1, 7):
                rank_check = promotion(self.char_data, self.career, self.upp, self.rank)
                if rank_check[1] == True:
                    self.rank = rank_check[0]
                    self.add_skill(self.skill_roll())
                else:
                    pass
            if str(self.rank) in self.char_data[self.career]["autoskills"]:
                if self.char_data[self.career]["autoskills"][str(self.rank)] in self.skills:
                    pass
                else:
                    self.add_skill(self.char_data[self.career]["autoskills"][str(self.rank)])
            self.aging()
            if self.terms < 7 and utility.dice(2, 6) >= self.char_data[self.career]["reenlist"] and survival(self.char_data, self.career, self.upp) == True:
                pass
            elif self.terms == 7 and utility.dice(2, 6) < 12:
                break
            elif self.terms == 8:
                break
            else:
                break
        self.mustering_out()
        characteristic_list_1 = ["STR1", "DEX1", "END1", "INT1", "EDU1", "SOC1"]
        characteristic_list_2 = ["STR2", "DEX2", "END2", "INT2", "EDU2", "SOC2"]
        characteristic_list_3 = ["STR-1", "DEX-1", "END-1", "INT-1", "EDU-1", "SOC-1"]
        for skill in self.skills.copy():
            if skill in characteristic_list_1:
                char_index = characteristic_list_1.index(skill)
                self.upp[char_index] += 1
                self.skills.pop(skill)
            if skill in characteristic_list_2:
                char_index = characteristic_list_2.index(skill)
                self.upp[char_index] += 2
                self.skills.pop(skill)
            if skill in characteristic_list_3:
                char_index = characteristic_list_3.index(skill)
                self.upp[char_index] -= 1
                self.skills.pop(skill)
        self.inventory_reorganizer()
        self.skill_stringer()
        self.rank_stringer()
        self.upp_stringer()
        self.inventory_stringer()
        self.char_stringer()

    # def career_choice(self):
    #     if max(self.upp) == self.upp[0]:
    #         self.career = "Marines"
    #     elif max(self.upp) == self.upp[1]:
    #         self.career = random.choice(["Army", "Scouts"])
    #     elif max(self.upp) == self.upp[3]:
    #         self.career = "Merchants"
    #     elif max(self.upp) == self.upp[4]:
    #         self.career = "Navy"
    #     else:
    #         self.career = random.choice(self.career_list)


    def add_skill(self, skill):
        if skill == "Blade Combat":
            skill = random.choice(blade_list)
        elif skill == "Gun Combat":
            skill = random.choice(gun_list)
        elif skill == "Vehicle":
            skill = random.choice(vehicle_list)
        if skill not in self.skills:
            self.skills[skill] = 1
        elif skill in self.skills:
            self.skills[skill] += 1
        else:
            pass


    def enlistment (self):
        enlist_roll = utility.dice(2, 6)
        if self.char_data[self.career]['enlistment+1'] == "":
            pass
        else:
            modifier_1 = stat_roll_parser(self.char_data[self.career]['enlistment+1'])
            modifier_2 = stat_roll_parser(self.char_data[self.career]['enlistment+2'])
            if int(self.upp[[modifier_1][0][0]]) >= int([modifier_1][0][1]):
                enlist_roll += 1
            if int(self.upp[[modifier_2][0][0]]) >= int([modifier_2][0][1]):
                enlist_roll += 2
            if enlist_roll >= self.char_data[self.career]["enlistment"]:
                pass
            else:
                self.career = random.choice(self.career_list)
            
    def skill_roll (self):
        if self.upp[4] >= 8:
            skill_table = random.choice (["personal", "service", "advanced", "advanced8"])
        else:
            skill_table = random.choice (["personal", "service", "advanced"])
        return random.choice(self.char_data[self.career][skill_table])


    def add_item(self, item, skills):
        item = inventory_check (item, skills)
        if item not in self.inventory:
            self.inventory[item] = 1
        elif item in self.inventory:
            self.inventory[item] += 1
        else:
            pass

    def mustering_out (self):
        characteristic_list_1 = ["STR1", "DEX1", "END1", "INT1", "EDU1", "SOC1"]
        characteristic_list_2 = ["STR2", "DEX2", "END2", "INT2", "EDU2", "SOC2"]
        cash_count = 0
        i = 0
        while i in range (0, self.terms+1):
            i += 1
            if utility.dice(1,6) <= 3 and cash_count <= 3:
                cash_count += 1
                cash_roll = utility.dice(1,6)
                if "Gambling" in self.skills:
                    cash_roll += 1
                self.cash += self.char_data[self.career]["cash"][cash_roll-1]
            else:
                material_roll = utility.dice(1,6)
                if self.rank >= 5:
                    material_roll += 1
                self.add_item(self.char_data[self.career]["benefits"][material_roll-1], self.skills)
                for item in self.inventory.copy():
                    if item in characteristic_list_1:
                        char_index = characteristic_list_1.index(item)
                        self.upp[char_index] += 1
                        self.inventory.pop(item)
                    elif item in characteristic_list_2:
                        char_index = characteristic_list_2.index(item)
                        self.upp[char_index] += 2
                        self.inventory.pop(item)

    
    def aging (self):
        if self.age >= 34 and self.age < 50:
            if utility.dice(2, 6) < 8:
                self.upp[0] -= 1
            if utility.dice(2, 6) < 7:
                self.upp[1] -= 1
            if utility.dice(2, 6) < 8:
                self.upp[2] -= 1
        elif self.age >= 50 and self.age < 66:
            if utility.dice(2, 6) < 9:
                self.upp[0] -= 1
            if utility.dice(2, 6) < 8:
                self.upp[1] -= 1
            if utility.dice(2, 6) < 9:
                self.upp[2] -= 1
        elif self.age >= 66:
            if utility.dice(2, 6) < 9:
                self.upp[0] -= 2
            if utility.dice(2, 6) < 9:
                self.upp[1] -= 2
            if utility.dice(2, 6) < 9:
                self.upp[2] -= 2
            if utility.dice(2, 6) < 9:
                self.upp[3] -= 1
        else:
            pass
        for characteristic in self.upp:
            if characteristic <= 0:
                if utility.dice(2, 6) < 8:
                    self.status = "DECEASED"
                else:
                    self.upp[characteristic] = 1
    

    def inventory_reorganizer(self):
        skill_list = list(self.skills.keys())
        skill_list.sort()
        inventory_temp = list(self.inventory.keys())
        used_weapon_list = []
        for skill in skill_list:
            if skill in gun_list or skill in blade_list:
                if skill not in inventory_temp:
                    for inventory_item in inventory_temp:
                        if inventory_item in gun_list or inventory_item in blade_list:
                            if skill not in used_weapon_list:
                                self.inventory[skill] = 1
                                used_weapon_list.append(skill)
                                self.inventory.pop(inventory_item)
                                inventory_temp.remove(inventory_item)
                            elif skill in used_weapon_list:
                                self.inventory[inventory_item] = inventory_item
                            else:
                                pass
                else:
                    pass
            else:
                pass


    def skill_stringer(self):
        for skill in self.skills:
            skill_description = skill + " " + str(self.skills[skill])
            if self.skill_string == "":
                self.skill_string = skill_description
            else:
                self.skill_string = ", ".join([self.skill_string, skill_description])
    
    def rank_stringer(self):
        self.rank_text = self.rank_text = self.char_data[self.career]["ranks"][str(self.rank)]
        if self.upp[5] == 11 and self.sex == "Male":
            self.rank_text += " Knight"
        elif self.upp[5] == 11 and self.sex == "Female":
            self.rank_text += " Dame"
        elif self.upp[5] == 12 and self.sex == "Male":
            self.rank_text += " Baron"
        elif self.upp[5] == 12 and self.sex == "Female":
            self.rank_text += " Baroness"
        elif self.upp[5] == 13 and self.sex == "Male":
            self.rank_text += " Marquis"
        elif self.upp[5] == 13 and self.sex == "Female":
            self.rank_text += " Marquesa"
        elif self.upp[5] == 14 and self.sex == "Male":
            self.rank_text += " Count"
        elif self.upp[5] == 14 and self.sex == "Female":
            self.rank_text += " Countess"
        elif self.upp[5] == 15 and self.sex == "Male":
            self.rank_text += " Duke"
        elif self.upp[5] == 15 and self.sex == "Female":
            self.rank_text += " Duchess"
        if "Medical" in self.skills:
            if self.skills["Medical"] >= 3:
               self.rank_text += " Dr."
            elif self.upp[4] >= 12 and self.upp[4] < 15:
                self.rank_text += " Dr."
            elif self.upp[4] >= 15:
                self.rank_text += " Prof."
        elif self.upp[4] >= 12 and self.upp[4] < 15:
                self.rank_text += " Dr."
        elif self.upp[4] >= 15:
                self.rank_text += " Prof."
        else:
            pass

    def upp_stringer(self):
        for digit in self.upp:
            if self.upp_string == "":
                self.upp_string = str(utility.pseudo_hex(digit))
            else:
                self.upp_string = "".join([self.upp_string, str(utility.pseudo_hex(digit))])

    def inventory_stringer(self):
        for item in self.inventory:
            final_item = item
            item_description = final_item + " x" + str(self.inventory[item])
            if self.inventory_string == "":
                self.inventory_string = item_description
            else:
                self.inventory_string = ", ".join([self.inventory_string, item_description])
    
    def char_stringer(self):
        self.char_string = f"{self.career} {self.rank_text} {self.name} {self.upp_string}   Age {self.age} {self.terms} terms Cr {self.cash}\n{self.species} {self.sex}\n{self.skill_string}\nInventory: {self.inventory_string}"


def get_maximum_cols(sheet):
    for i in range(1, 20000):
        if sheet.cell(row=2, column= i).value == None:
            max_col = i
            break
    return max_col

def get_maximum_rows(sheet):
    for i in range(1, 20000):
        if sheet.cell(row=i, column = 1).value == None:
            max_row = i 
            break
    return max_row

def output_to_excel(filename, character):
    full_filename = str(filename +".xlsx")
    if not utility.check_file_exists(full_filename, ):
        workbook = openpyxl.Workbook()
        sheet_object = workbook.active
        sheet_object.title = "Characters"
        sheet_object['A1'] = "Classic Traveller Characters"
        sheet_object['A1'].font = Font(size=16, bold=True)
        for cell in ["A2", "B2", "C2", "D2", "E2", "F2", "G2", "H2", "I2", "J2", "K2", "L2", "M2", "N2", "O2", "P2"]:
            sheet_object[cell].font = Font(bold=True)
        sheet_object["A2"] = "Name"
        sheet_object["B2"] = "Service"
        sheet_object["C2"] = "Rank"
        sheet_object["D2"] = "STR"
        sheet_object["E2"] = "DEX"
        sheet_object["F2"] = "END"
        sheet_object["G2"] = "INT"
        sheet_object["H2"] = "EDU"
        sheet_object["I2"] = "SOC"
        sheet_object["J2"] = "Age"
        sheet_object["K2"] = "Terms"
        sheet_object["L2"] = "Credits"
        sheet_object["M2"] = "Species"
        sheet_object["N2"] = "Sex"
        sheet_object["O2"] = "Skills"
        sheet_object["P2"] = "Inventory"
    else:
        workbook = openpyxl.load_workbook(full_filename)
    sheet_object = workbook.active
    current_row = get_maximum_rows(sheet_object)
    sheet_object.cell(row=current_row, column=1).value = str(character.name)
    sheet_object.cell(row=current_row, column=2).value = str(character.career)
    sheet_object.cell(row=current_row, column=3).value = str(character.rank_text)
    sheet_object.cell(row=current_row, column=4).value = str(character.upp[0])
    sheet_object.cell(row=current_row, column=5).value = str(character.upp[1])
    sheet_object.cell(row=current_row, column=6).value = str(character.upp[2])
    sheet_object.cell(row=current_row, column=7).value = str(character.upp[3])
    sheet_object.cell(row=current_row, column=8).value = str(character.upp[4])
    sheet_object.cell(row=current_row, column=9).value = str(character.upp[5])
    sheet_object.cell(row=current_row, column=10).value = str(character.age)
    sheet_object.cell(row=current_row, column=11).value = str(character.terms)
    sheet_object.cell(row=current_row, column=12).value = str(character.cash)
    sheet_object.cell(row=current_row, column=13).value = str(character.species)
    sheet_object.cell(row=current_row, column=14).value = str(character.sex)
    sheet_object.cell(row=current_row, column=15).value = str(character.skill_string)
    sheet_object.cell(row=current_row, column=16).value = str(character.inventory_string)
    workbook.save(full_filename)


if __name__ == "__main__":
    # utility.clear_screen()
    # career_choice = utility.menu("Choose Your Career", "1. Army", "2. Navy", "3. Scouts", "4. Marines", "5. Merchants", "6. Other")
    # if career_choice == 1:
    #     career = "Army"
    # elif career_choice == 2:
    #     career = "Navy"
    # elif career_choice == 3:
    #     career = "Scouts"
    # elif career_choice == 4:
    #     career = "Marines"
    # elif career_choice == 5:
    #     career = "Merchants"
    # elif career_choice == 6:
    #     career = "Other"  
    # print("-------------------------------")
    input_loop = True
    while input_loop:
        career = str(input("Please input your career (Army, Navy, Scouts, Marines, Merchants, or Other):"))
        if career in ("Army", "Navy", "Scouts", "Marines", "Merchants", "Other"):
            break
        else:
            print ("Invalid Career")
    character_number = int(input("Please input how many characters you want (numbers only):"))
    for i in range (0, character_number+1):
        test = Character("Data/careers.json", False, career)
        output_to_excel("test_file", test)